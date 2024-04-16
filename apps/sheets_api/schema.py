import graphene
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Assume the workbook is stored at this location
WORKBOOK_LOCATION = "sheets_api/templates/The_Statutory_Biodiversity_Metric_Calculation_Tool_-_Macro_disabled_02.24.xlsx"

class Cell(graphene.ObjectType):
    row = graphene.Int(required=True)
    column = graphene.String(required=True)
    value = graphene.String()

class CellInput(graphene.InputObjectType):
    row = graphene.Int(required=True)
    column = graphene.String(required=True)
    value = graphene.String()

class Query(graphene.ObjectType):
    cell = graphene.Field(Cell, sheet=graphene.String(required=True), row=graphene.Int(required=True), column=graphene.String(required=True))

    def resolve_cell(self, info, sheet, row, column):
        wb = load_workbook(WORKBOOK_LOCATION)
        ws = wb[sheet]
        column_index = column_index_from_string(column)
        cell = ws.cell(row=row, column=column_index)
        return Cell(row=row, column=column, value=str(cell.value))

class UpdateCell(graphene.Mutation):
    class Arguments:
        sheet = graphene.String(required=True)
        cell_input = CellInput(required=True)

    cell = graphene.Field(Cell)

    def mutate(self, info, sheet, cell_input):
        wb = load_workbook(WORKBOOK_LOCATION)
        ws = wb[sheet]
        column_index = column_index_from_string(cell_input.column)
        cell = ws.cell(row=cell_input.row, column=column_index)
        cell.value = cell_input.value
        wb.save(WORKBOOK_LOCATION)
        return UpdateCell(cell=Cell(row=cell_input.row, column=cell_input.column, value=cell_input.value))

class Mutation(graphene.ObjectType):
    update_cell = UpdateCell.Field()

schema = graphene.Schema(query=Query, mutation=Mutation)

# querying multiple cells
# mutating multiple cells
# lazy evaluation of a list of commands - transactions
# Security of editing a file on disk?
# Copy and lazy evaluate commands against template file?
# Find columns by name
# Model by xlsx which stores the template location
# Package this
# Write readme
# Link to recipe from form somehow? Use service for recipe logic
# Add auth
